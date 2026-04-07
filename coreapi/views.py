from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.hashers import check_password
from .models import UserProfile, ReportSketch
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.middleware.csrf import get_token
from django_ratelimit.decorators import ratelimit
import os, io
import mimetypes
from django.http import JsonResponse, FileResponse, Http404
from django.views.decorators.http import require_http_methods, require_POST, require_GET
from PyPDF2 import PdfMerger
from PIL import Image, ImageDraw
from datetime import datetime, timedelta
import time
from django.conf import settings
from coreapi.search_index import get_index
from coreapi.search_index import refresh_index
from docx import Document
import openpyxl
from django.views.decorators.csrf import ensure_csrf_cookie
from django.core.files.base import ContentFile
from django.template.loader import render_to_string
import base64
from django.db.models import Q, Max
import uuid
from django.contrib.auth.decorators import login_required
from weasyprint import HTML, CSS
from .models import UserProfile, SiteVisitReport
import fitz  
from django.shortcuts import render, get_object_or_404
import shutil
from urllib.parse import unquote
from django.http import HttpResponseNotFound, HttpResponseBadRequest
import re
from django.utils import timezone
from chat.models import FolderChatMessage, FolderChatVisit 
from django.core.cache import cache
from .utils import get_case_folder_info
from playwright.sync_api import sync_playwright
import sys
import asyncio
from .models import ClientFolder
from django.db import transaction
from django.core.files.storage import FileSystemStorage
import psutil
import logging
import subprocess
import shlex
from django.views.decorators.csrf import csrf_exempt

import fitz  # PyMuPDF
from django.http import HttpResponse, HttpResponseNotFound
from django.core.cache import cache


# Security Check: Only allow superusers (Developers/Admins)
def is_developer(user):
    return user.is_authenticated and user.is_superuser


# Grab the logger so we can write to it
logger = logging.getLogger('coreapi')


def get_system_logs(line_count=500):
    terminal_log_path = os.path.join(settings.BASE_DIR, 'logs', 'terminal.log')
    
    if not os.path.exists(terminal_log_path):
        return "System initializing... Waiting for terminal output."
    
    try:
        # Read as RAW BYTES to bypass Windows encoding corruption
        with open(terminal_log_path, 'rb') as f:
            raw_bytes = f.read()
            
        # Detect if PowerShell saved it as UTF-16 (spacy text) or standard UTF-8
        if b'\x00' in raw_bytes:
            raw_text = raw_bytes.decode('utf-16le', errors='ignore')
        else:
            raw_text = raw_bytes.decode('utf-8', errors='ignore')
            
        # Strip all null bytes and crush multiple blank lines into one
        raw_text = raw_text.replace('\x00', '')
        clean_lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
        
        return "\n".join(clean_lines[-line_count:])
            
    except Exception as e:
        return f"Error reading terminal log: {str(e)}"

    
# --- UPDATE YOUR EXISTING DASHBOARD VIEW ---
def developer_dashboard(request):
    if request.session.get("user_name") != 'alnroy':
        return redirect('coreapi:login_page')

    # Force a log entry so you see it working immediately!
    logger.info(f"Admin '{request.session.get('user_name')}' accessed the Command Center.")

    cpu_usage = psutil.cpu_percent(interval=0.1)
    memory = psutil.virtual_memory()
    ram_usage = memory.percent
    ram_total = round(memory.total / (1024 ** 3), 1)

    active_users = []
    all_users = UserProfile.objects.all()
    for user in all_users:
        if cache.get(f"online_user_{user.user_name}"):
            active_users.append(user)
    
    context = {
        'cpu_usage': cpu_usage,
        'ram_usage': ram_usage,
        'ram_total': ram_total,
        'active_users': active_users,

        'system_logs': get_system_logs(50),  # Pass initial logs to the template
    }
    return render(request, 'dev_dashboard.html', context)


def get_latest_error_api(request):
    log_path = os.path.join(settings.BASE_DIR, 'logs', 'latest_error.json')
    if os.path.exists(log_path):
        with open(log_path, 'r') as f:
            return JsonResponse(json.load(f))
    return JsonResponse({'type': 'None'})


@csrf_exempt
def clear_stale_sessions_api(request):
    if request.session.get("user_name") != 'alnroy':
        return JsonResponse({'error': 'Unauthorized'}, status=401)
        
    if request.method == 'POST':
        # Loop through users and delete their specific online cache keys
        all_users = UserProfile.objects.all()
        for user in all_users:
            cache.delete(f"online_user_{user.user_name}")
            
        return JsonResponse({'success': True, 'message': 'Stale sessions cleared!'})
        
    return JsonResponse({'error': 'Invalid request'}, status=400)


# --- ADD THIS NEW API FOR THE REFRESH BUTTON ---
def fetch_live_logs_api(request):
    """API to fetch logs without reloading the page"""
    if request.session.get("user_name") != 'alnroy':
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    logs = get_system_logs(100)  # Fetch more lines on refresh
    return JsonResponse({'logs': logs})


@csrf_exempt
def execute_command_api(request):
    if request.session.get("user_name") != 'alnroy':
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            command = data.get('command', '').strip()
            
            if not command:
                return JsonResponse({'output': ''})

            # Log the command so you have an audit trail of what you did
            logger.info(f"TERMINAL EXECUTION by alnroy: {command}")

            # Run the command in the actual OS shell. 
            # Timeout is 15 seconds so you don't accidentally freeze the server with an infinite loop.
            result = subprocess.run(
                command,
                shell=True,
                capture_output=True,
                text=True,
                timeout=15
            )
            
            # Combine standard output and error output
            output = result.stdout + result.stderr
            return JsonResponse({'output': output})
            
        except subprocess.TimeoutExpired:
            return JsonResponse({'output': '\nError: Command timed out after 15 seconds.'})
        except Exception as e:
            return JsonResponse({'output': f'\nSystem Error: {str(e)}'})

    return JsonResponse({'error': 'Invalid request'}, status=400)


@csrf_protect
def refresh_files(request):
    try:
        # Rebuild the index synchronously
        refresh_index()
        return JsonResponse({"status": "success", "message": "Index refreshed"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)
# ----------------------------
# Login / Logout / Dashboard
# ----------------------------


@ensure_csrf_cookie
def login_page(request):
    if request.session.get("user_id"):
        return redirect("coreapi:dashboard")
    return render(request, "login.html")


def dashboard(request):
    role = request.session.get("user_role")
    
    # 1. Admin goes to the new Core Dashboard
    if role in ["admin", "accountant"]:
        return redirect("core:admin_dashboard") 
        
    # 2. Others go to the Office Dashboard
    elif role == "site":
        return redirect("coreapi:office_dashboard")
    
    elif role == "office":
        return redirect("coreapi:office")
    
    elif role == "IT":
        return redirect("coreapi:dev_dashboard")
    
    # 3. Unknown roles get rejected
    return JsonResponse({"error": "Invalid role"}, status=500)
# ==========================


def server_health_api(request):
    if request.session.get("user_name") != 'alnroy':
        return JsonResponse({'error': 'Unauthorized'}, status=401)
        
    memory = psutil.virtual_memory()
    
    return JsonResponse({
        'cpu_usage': psutil.cpu_percent(interval=0.1),
        'ram_usage': memory.percent,
        'ram_total': round(memory.total / (1024 ** 3), 1)
    })

    
# --- 2. RESTART SERVER API ---
@csrf_exempt
def restart_server_api(request):
    if request.session.get("user_name") != 'alnroy':
        return JsonResponse({'error': 'Unauthorized'}, status=401)
        
    if request.method == 'POST':
        try:
            # The Hack: Touch manage.py to trigger Django's StatReloader
            manage_py_path = os.path.join(settings.BASE_DIR, 'manage.py')
            os.utime(manage_py_path, None) 
            
            # Log the restart so it appears in your web terminal
            logger.info("SYSTEM ALERT: Administrator triggered a manual server restart.")
            
            return JsonResponse({'success': True, 'message': 'Restart signal sent. Server is rebooting...'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
            
    return JsonResponse({'error': 'Invalid method'}, status=400)


@csrf_protect
def office_verification(request):
    """
    Renders the three-part office verification dashboard.
    """
    if request.session.get("user_role") not in ["office", "IT"]:
        return redirect("coreapi:login_page")
        
    rel_path = request.GET.get("path", "").strip("/")
    report = None
    
    if rel_path:
        # Fetch the site staff's report for this folder
        report = SiteVisitReport.objects.filter(target_folder=rel_path).first()

    context = {
        'current_path': rel_path,
        'report': report,
        'site_data': report.form_data if report else {},
        'google_maps_api_key': settings.GOOGLE_MAPS_API_KEY,
    }
    return render(request, "office_verification.html", context)


@csrf_protect
def get_site_report_data(request):
    file_no = request.GET.get('file_no')
    folder_path = request.GET.get('path', '')
    
    report = None

    # 1. Search Logic for Site Report
    if raw_path := request.GET.get('path', ''):
        clean_path = unquote(raw_path).replace('\\', '/').replace('G:/My Drive/', '').strip('/')
        report = SiteVisitReport.objects.filter(target_folder__endswith=clean_path).order_by('-updated_at').first()

    if not report and file_no:
        report = SiteVisitReport.objects.filter(office_file_no=file_no).first()

    if report:
        # Get raw site data
        site_data = report.form_data or {}
        
        # --- NEW: GET VERIFICATION DATA ---
        verification_data = {}
        try:
            # Try to find an existing Verification Report
            v_report = VerificationReport.objects.get(office_file_no=report.office_file_no)
            verification_data = v_report.verification_database
            
            # If the admin already edited the documents list or inspection date, 
            # push it into the payload so the frontend gets the latest version
            if v_report.documents_received:
                if 'Valuers_Checklist' not in site_data:
                    site_data['Valuers_Checklist'] = {}
                site_data['Valuers_Checklist']['documents_received'] = v_report.documents_received
                
            if v_report.inspection_date:
                if 'Valuers_Checklist' not in site_data:
                    site_data['Valuers_Checklist'] = {}
                site_data['Valuers_Checklist']['inspection_date'] = v_report.inspection_date
                
        except VerificationReport.DoesNotExist:
            pass  # No verification done yet, which is totally fine

        # Store original site visit data in a dedicated reference key
        # AND purge the main DynamicDocuments from the payload to avoid auto-populating cards
        original_docs = site_data.get('DynamicDocuments', [])
        site_data['OriginalSiteDocuments'] = original_docs
        site_data['DynamicDocuments'] = [] 

        # Inject the verification database safely into the payload
        if verification_data:
            site_data['DynamicDocuments'] = verification_data.get('DynamicDocuments', [])
            site_data['BoundaryAnalysisDocs'] = verification_data.get('BoundaryAnalysisDocs', [])
            
            # Map database keys to frontend expected keys
            site_data['owners_data'] = verification_data.get('OwnersData', [])
            site_data['schedule_data'] = verification_data.get('ScheduleData', {})
            site_data['survey_land_extend'] = verification_data.get('SurveyAnalysis', {})
            site_data['survey_notes'] = verification_data.get('SurveyNotes', '')
            
            # Update Valuers_Checklist for top summary strip
            if 'Valuers_Checklist' not in site_data:
                site_data['Valuers_Checklist'] = {}
            
            site_data['Valuers_Checklist']['applicant_name'] = verification_data.get('ApplicantName', '')
            site_data['Valuers_Checklist']['product'] = verification_data.get('Product', '')
            site_data['Valuers_Checklist']['person_met'] = verification_data.get('PersonMet', '')
            
            # Keep legacy key for compatibility during transition if needed
            site_data['Verification_Database'] = verification_data

        # 2. Prepare Metadata for Sidebar
        meta = {
            "user": report.user.user_name if report.user else "Unknown",
            "office_file_no": report.office_file_no,
            "applicant_name": report.applicant_name,
            "target_folder": report.target_folder,
            "created_at": report.created_at.strftime("%d-%b-%Y %I:%M %p"),
            "generated_pdf_name": report.generated_pdf_name or "Not Generated",
            "completion_score": report.completion_score
        }
        
        return JsonResponse({
            'found': True,
            'data': site_data,
            'meta': meta,
            'real_file_no': report.office_file_no
        })
    else:
        return JsonResponse({'found': False, 'message': 'Report not found'})

      
def save_office_corrections(request):
    """
    Saves corrections back to the SiteVisitReport form_data.
    Handles nested keys like 'Valuers_Checklist.applicant_name'
    """
    if request.method == "POST":
        try:
            payload = json.loads(request.body)
            file_no = payload.get('file_no')
            corrections = payload.get('corrections', {})  # e.g. {'Valuers_Checklist.applicant_name': 'New Name'}

            if not file_no:
                return JsonResponse({'error': 'Missing file number'}, status=400)

            report = SiteVisitReport.objects.filter(office_file_no=file_no).first()
            if not report:
                return JsonResponse({'error': 'Report not found'}, status=404)

            # Get current data
            current_data = report.form_data if report.form_data else {}
            
            # Apply corrections using dot notation (Parent.Child)
            for path, value in corrections.items():
                keys = path.split('.')
                d = current_data
                # Navigate to the last key
                for key in keys[:-1]:
                    if key not in d: 
                        d[key] = {}  # Create dict if missing
                    d = d[key]
                # Set the value
                d[keys[-1]] = value
            
            # Save back to DB
            report.form_data = current_data
            report.save()

            return JsonResponse({'success': True, 'message': 'Corrections saved successfully'})

        except Exception as e:
            print(f"Error saving: {e}")
            return JsonResponse({'error': str(e)}, status=500)
            
    return JsonResponse({'error': 'Invalid method'}, status=405)


@csrf_protect
@ratelimit(key="ip", rate="5/m", block=True)
def login_api(request):
    get_token(request)

    if request.method != "POST":
        return JsonResponse({"error": "POST request required"}, status=400)

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    email = data.get("email", "").strip()
    password = data.get("password", "")

    if not email or not password:
        return JsonResponse({"error": "Email and password required"}, status=400)

    try:
        user = UserProfile.objects.get(email=email)
    except UserProfile.DoesNotExist:
        return JsonResponse({"error": "Invalid email or password"}, status=401)

    if not check_password(password, user.password):
        return JsonResponse({"error": "Invalid email or password"}, status=401)

    # SESSION SAVE
    request.session["user_id"] = user.id
    request.session["user_name"] = user.user_name
    request.session["user_email"] = user.email
    request.session["user_role"] = user.role
    request.session.set_expiry(604800)  
    request.session.modified = True

    # 🚨 THE FIX: Intercept the developer login
    if user.user_name == 'alnroy':
        redirect_target = "/coreapi/dev-center/"  # URL to your Dev Dashboard
    else:
        redirect_target = "/coreapi/dashboard/"  # URL for everyone else

    return JsonResponse({
        "success": True,
        "message": "Login successful",
        "redirect": redirect_target
    })


def logout_api(request):
    request.session.flush()
    return redirect("coreapi:login_page")


# ----------------------------
# File / Folder Handling
# ----------------------------
DOCUMENTS_FOLDER = settings.DOCUMENTS_ROOT


@csrf_protect
def office(request):
    """
    Renders the Office Landing Page with Actions & Profile.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("coreapi:login_page")
        
    try:
        user_profile = UserProfile.objects.get(id=user_id)
    except UserProfile.DoesNotExist:
        return redirect("coreapi:login_page")

    context = {
        "user_profile": user_profile,
        "last_login": user_profile.last_seen if user_profile.last_seen else "Never",
        "created_at": user_profile.created_at
    }
    return render(request, "office.html", context)


def create_folder_page(request):
    return render(request, "create_folder.html")


@csrf_protect
@require_POST
def create_folder_api(request):
    try:
        user_id = request.session.get("user_id")
        if not user_id: return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=401)

        data = json.loads(request.body)
        
        # 1. Extract Data
        structure = data.get('structure', {})
        meta = data.get('metadata', {})
        
        bank_code = meta.get('bank_code') 
        dist_code = meta.get('dist_code') 
        year = meta.get('year', '26')    
        
        # 3. DB Transaction (Get Next Sequence)
        with transaction.atomic():
            # USE MAX SHIFT: prevent collision if a previous file was deleted
            res = ClientFolder.objects.filter(
                year=year,
                bank_code=bank_code,
                district_code=dist_code
            ).aggregate(Max('sequence_no'))
            
            max_seq = res.get('sequence_no__max') or 0
            next_seq = max_seq + 1
            seq_str = f"{next_seq:04d}"
            
            # 4. Generate Unique ID & Name
            unique_id = f"{bank_code}{year}{dist_code}{seq_str}"
            applicant_clean = meta.get('applicant', 'XXX')
            
            # --- DETERMINE WHICH EXTRA STRING TO USE ---
            local_body = meta.get('local_body', '')
            muthoot_branch = meta.get('muthoot_branch', '')
            
            extra_str = ""
            if local_body:
                extra_str = f"_{local_body}"
            elif muthoot_branch:
                extra_str = f"_{muthoot_branch}"
                
            # --- BUILD FOLDER NAME ---
            folder_name = f"{unique_id}_#{applicant_clean}#_{meta.get('product')}_{meta.get('dist_name')}{extra_str}_{meta.get('date_str')}_{meta.get('site_code')}_{meta.get('off_code')}_{meta.get('bank_ref')}"
            # ---------------------------------------------------------
            # 🔧 PATH FIX: Force Root to 'G:\My Drive'
            # ---------------------------------------------------------
            #  'settings.DOCUMENTS_ROOT' is 'G:\My Drive\2026_2027'
            # So we go one level up to get 'G:\My Drive'
            
            base_drive_root = os.path.dirname(settings.DOCUMENTS_ROOT) 
            # OR explicitly: base_drive_root = r"G:\My Drive"
            
            # Construct: G:\My Drive \ 2026_2027 \ 1000.HDFC \ KL01.TVM \ Folder...
            path_components = [
                base_drive_root,
                structure.get('year', '2026_2027'),
                structure.get('bank_folder', 'Unknown'),
                structure.get('dist_folder', 'Unknown'),
                folder_name
            ]
            full_path = os.path.join(*path_components)
            full_path = os.path.normpath(full_path)
            
            if os.path.exists(full_path):
                return JsonResponse({'success': False, 'error': 'System Error: Folder exists on disk but not in DB.'})

            # 5. Create Folders
            os.makedirs(full_path, exist_ok=True)

            # 6. Save to DB
            ClientFolder.objects.create(
                unique_file_no=unique_id,
                year=year,
                bank_code=bank_code,
                district_code=dist_code,
                sequence_no=next_seq,
                applicant_name=meta.get('applicant'),
                product=meta.get('product'),
                bank_ref_no=meta.get('bank_ref'),
                site_staff_code=meta.get('site_code'),
                office_staff_code=meta.get('off_code'),
                full_folder_path=full_path
            )

            # --- NEW: COPY EXCEL TEMPLATE ---
            try:
                # 1. Get Bank Name (from our global list)
                bank_map = {b['code']: b['name'] for b in BANKS}
                bank_name = bank_map.get(bank_code)

                if bank_name:
                    # 2. Check for template in static/excels/
                    template_filename = f"{bank_name}.xlsm"
                    template_src = os.path.join(settings.BASE_DIR, 'static', 'excels', template_filename)

                    if os.path.exists(template_src):
                        # 3. New Format: OfficeFileNo_BankFileNo_Applicant
                        office_no = unique_id
                        bank_ref = str(meta.get('bank_ref', 'XXXX')).strip().replace(' ', '_')
                        applicant_sanitized = str(meta.get('applicant', 'Report')).strip().replace(' ', '_')
                        
                        new_excel_name = f"{office_no}_{bank_ref}_{applicant_sanitized}.xlsm"
                        excel_dest = os.path.join(full_path, new_excel_name)

                        # 4. Copy the file
                        shutil.copy2(template_src, excel_dest)
                        print(f"Excel template copied: {new_excel_name}")
            except Exception as ex:
                # We skip errors here so the folder creation isn't aborted if template copy fails
                print(f"Warning: Excel template copy skipped: {ex}")

        return JsonResponse({'success': True, 'new_id': unique_id})

    except Exception as e:
        print(f"Error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
        

@require_GET
def get_next_sequence_api(request):
    # 1. Get parameters from the frontend
    bank_code = request.GET.get('bank_code')
    dist_code = request.GET.get('dist_code')
    year = request.GET.get('year', '26')

    # 2. Safety Check
    if not (bank_code and dist_code):
        return JsonResponse({'seq': '0001'})

    # 3. Count existing files for this SPECIFIC combo
    # e.g. Count HDFC (01) files in TVM (01)
    count = ClientFolder.objects.filter(
        year=year,
        bank_code=bank_code,
        district_code=dist_code
    ).count()

    # 4. Return next number (Count + 1)
    next_seq = count + 1
    return JsonResponse({'seq': f"{next_seq:04d}"})  # Returns "0043"


@require_GET
def check_duplicate_api(request):
    bank_code = request.GET.get('bank_code')
    bank_ref = request.GET.get('bank_ref')
    
    if not bank_code or not bank_ref:
        return JsonResponse({'exists': False})

    # Check if this Bank Ref exists for this specific Bank
    exists = ClientFolder.objects.filter(
        bank_code=bank_code,
        bank_ref_no=bank_ref
    ).exists()
    
    return JsonResponse({'exists': exists})


@require_POST
def upload_site_photos_api(request):
    """
    Uploads multiple images to a 'P' subfolder inside the target directory.
    """
    if request.FILES:
        try:
            # 1. Get the current working folder path from the frontend
            target_folder_rel = request.POST.get('current_folder', '')
            
            if not target_folder_rel:
                return JsonResponse({'success': False, 'error': 'No folder selected'})

            # 2. Construct absolute path: G:\My Drive\...\CurrentFolder\p
            # We assume settings.DOCUMENTS_ROOT is your base storage path
            base_path = settings.DOCUMENTS_ROOT 
            target_folder_abs = os.path.join(base_path, target_folder_rel, 'P')

            # 3. Create 'p' folder if it doesn't exist
            if not os.path.exists(target_folder_abs):
                os.makedirs(target_folder_abs, exist_ok=True)

            saved_files = []
            fs = FileSystemStorage(location=target_folder_abs)

            # 4. Loop through uploaded files and save them
            files = request.FILES.getlist('site_photos')
            for f in files:
                # Save file (handles duplicate names automatically)
                filename = fs.save(f.name, f)
                
                # Store relative path for the database/frontend
                # Format: CurrentFolder/p/filename.jpg
                rel_path = os.path.join(target_folder_rel, 'P', filename).replace('\\', '/')
                saved_files.append(rel_path)

            return JsonResponse({'success': True, 'saved_paths': saved_files})

        except Exception as e:
            print(f"Upload Error: {e}")
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'No files provided'})


def office_dashboard(request):
    """
    Loads the dashboard instantly. 
    Only fetches the top-level folders to populate the sidebar/initial view.
    No files are loaded yet.
    """
    # Only scan the immediate root directory, do NOT walk the whole tree
    folders = scan_root_folders_only(DOCUMENTS_FOLDER) 
    
    context = {
        "folders": folders,
        "files": [],  # Empty on purpose. JS will fetch files when a user clicks a folder.
    }
    return render(request, "office_dashboard.html", context)


def scan_root_folders_only(base_folder):
    """
    LIGHTWEIGHT SCANNER: Uses os.scandir for maximum speed.
    Only returns folders in the immediate root.
    """
    folders = []
    try:
        # os.scandir is significantly faster than os.listdir or os.walk
        with os.scandir(base_folder) as it:
            for entry in it:
                if entry.is_dir():
                    folders.append({
                        "id": entry.name,
                        "name": entry.name,
                        "path": entry.name,  # Relative path at root is just the name
                        "type": "folder"
                    })
    except Exception as e:
        print(f"Error scanning root: {e}")
        
    # Sort alphabetically
    return sorted(folders, key=lambda x: x['name'].lower())


def categorize_file(extension):
    categories = {
        'A': ['.pdf', '.doc', '.docx'],
        'B': ['.xls', '.xlsx', '.csv'],
        'C': ['.jpg', '.jpeg', '.png', '.gif', '.bmp'],
    }
    for cat, exts in categories.items():
        if extension in exts:
            return cat
    return 'Other'


def get_file_type_description(extension):
    types = {
        '.pdf': 'PDF Document',
        '.doc': 'Word Document',
        '.docx': 'Word Document',
        '.xls': 'Excel Spreadsheet',
        '.xlsx': 'Excel Spreadsheet',
        '.csv': 'CSV File',
        '.jpg': 'Image',
        '.jpeg': 'Image',
        '.png': 'Image',
        '.gif': 'Image',
        '.txt': 'Text File',
    }
    return types.get(extension, 'Document')


def format_file_size(size_bytes):
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.1f} TB"


# ----------------------------
# APIs
# ----------------------------
@require_http_methods(["GET"])
def search_folders_api(request):
    q = request.GET.get("q", "").lower()
    
    # 1. Get from Index instead of Disk
    index = get_index()
    all_folders = index.get("folders", [])

    # 2. Filter
    if q:
        folders = [
            f for f in all_folders 
            if q in f["name"].lower() or q in f.get("id", "").lower()
        ]
    else:
        # If no query, return top-level folders (or first 50 cached folders)
        # Assuming your index stores root paths correctly
        folders = all_folders[:50] 

    # 3. Sort Newest First
    folders.sort(key=lambda x: x.get('mtime', 0), reverse=True)

    return JsonResponse({"folders": folders})


# search files-----------------
def search_files(request):
    q = request.GET.get("q", "").strip().lower()
    
    if not q or len(q) < 2:
        return JsonResponse({"folders": [], "files": []})

    # 1. Get the Index (Instant RAM access)
    index = get_index()
    
    # --- HELPER: Sort by Modified Time (Descending) ---
    # Requires 'mtime' to be saved in your index. If missing, defaults to 0.
    def sort_newest(item):
        return item.get('mtime', 0)

    # 2. Filter & Sort FOLDERS
    # We filter first, THEN sort, THEN slice.
    matched_folders = [
        f for f in index.get("folders", []) 
        if q in f["name"].lower()
    ]
    # Sort: Newest First
    matched_folders.sort(key=sort_newest, reverse=True)
    # Slice: Take top 20
    matched_folders = matched_folders[:20]

    # --- PROCESS FOLDERS (Chat/Unread Status) ---
    processed_folders = []
    
    user_id = request.session.get("user_id")
    current_user_profile = None
    if user_id:
        try: current_user_profile = UserProfile.objects.get(id=user_id)
        except: pass

    for folder in matched_folders:
        # Create a copy so we don't mutate the cached index
        f_copy = folder.copy()
        
        name = f_copy['name']
        path = f_copy['path']
        
        has_hash = "#" in name
        is_case_folder = re.search(r'^\d+_.*\d{2}\.\d{2}\.\d{4}', name)
        has_chat = bool(has_hash or is_case_folder)
        
        is_unread = False
        if has_chat and current_user_profile:
            is_unread = check_unread_status(current_user_profile, path)

        f_copy['has_chat'] = has_chat
        f_copy['is_unread'] = is_unread
        processed_folders.append(f_copy)

    # 3. Filter & Sort FILES
    matched_files = [
        f for f in index.get("files", []) 
        if q in f["name"].lower()
    ]
    # Sort: Newest First
    matched_files.sort(key=sort_newest, reverse=True)
    # Slice: Take top 50
    matched_files = matched_files[:50]

    return JsonResponse({
        "folders": processed_folders,
        "files": matched_files
    })


@require_http_methods(["GET"])
def get_folder_contents_api(request):
    rel_path = request.GET.get("path", "").strip("/")
    recursive = request.GET.get("recursive", "0") == "1"
    page = int(request.GET.get("page", 1))
    limit = int(request.GET.get("limit", 50))
    
    base = os.path.realpath(DOCUMENTS_FOLDER)
    abs_path = os.path.realpath(os.path.join(base, rel_path))
    
    saved_draft = None
    folder_info = None
    auto_fill_data = None
    vetting_data = None

    if not abs_path.startswith(base) or not os.path.isdir(abs_path):
        return JsonResponse({"folders": [], "files": [], "has_next": False})

    # --- METADATA & DRAFT FETCHING (Page 1 Only) ---
    if page == 1:
        if rel_path:
            current_folder_name = os.path.basename(abs_path)
            # 1. Parse the folder name to get the exact file number
            auto_fill_data = parse_folder_metadata(current_folder_name)
            file_no = auto_fill_data.get('file_no')
            
            # 2. Search Database by File Number (NOT by folder path)
            if file_no:
                draft = SiteVisitReport.objects.filter(office_file_no=file_no).order_by('-updated_at').first()
                if draft and draft.form_data:
                    saved_draft = draft.form_data if isinstance(draft.form_data, dict) else json.loads(draft.form_data)
                    saved_draft['report_id'] = draft.id
                    sketches = ReportSketch.objects.filter(report=draft)
                    if 'images' not in saved_draft: 
                        saved_draft['images'] = {}
                    for sketch in sketches:
                        if sketch.image: 
                            saved_draft['images'][sketch.source_key] = sketch.image.url
                            
        folder_name = os.path.basename(abs_path)
        if "#" in folder_name or re.search(r'^\d+_.*\d{2}\.\d{2}\.\d{4}', folder_name):
            # Fetch the creation date from the ClientFolder table using the file number
            case_folder = ClientFolder.objects.filter(unique_file_no=file_no).first() if file_no else None
            folder_info = get_case_folder_info(abs_path, db_created_at=case_folder.created_at if case_folder else None)

    # --- SCANNING & SORTING ---
    all_folders = []
    all_files = []
    
    user_id = request.session.get("user_id")
    current_user_profile = None
    if user_id:
        try: current_user_profile = UserProfile.objects.get(id=user_id)
        except: pass

    try:
        if recursive:
            # Flattened recursive file search
            for root, dirs, files in os.walk(abs_path):
                for name in files:
                    if name.lower() in ['desktop.ini', '.ds_store']: continue
                    full_abs = os.path.join(root, name)
                    rel_to_base = os.path.relpath(full_abs, base).replace('\\', '/')
                    
                    # Create a simple object mirroring the os.DirEntry interface
                    class MockEntry:
                        def __init__(self, n, p, rp):
                            self.name = n
                            self.path = p
                            self.rel_path = rp
                            self._stat = None
                        def is_dir(self): return False
                        def stat(self):
                            if not self._stat: self._stat = os.stat(self.path)
                            return self._stat
                    
                    all_files.append(MockEntry(name, full_abs, rel_to_base))
            all_folders = [] # Return no folders when recursive
        else:
            with os.scandir(abs_path) as it:
                for entry in it:
                    if entry.is_dir():
                        all_folders.append(entry)
                    else:
                        all_files.append(entry)

        # Sort by MTime Descending
        if all_folders:
            all_folders.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        all_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)

        # --- PAGINATION ---
        start = (page - 1) * limit
        end = start + limit
        
        folders_to_process = all_folders[start:end] if not recursive else []
        # If we didn't fill the limit with folders, fill with files
        remaining_limit = limit - len(folders_to_process)
        file_start = max(0, start - len(all_folders))
        files_to_process = all_files[file_start : file_start + remaining_limit]

        has_next = (len(all_folders) + len(all_files)) > end

        # --- PROCESS FOLDERS ---
        final_folders = []
        for entry in folders_to_process:
            folder_stats = entry.stat()
            name = entry.name
            full_rel_path = f"{rel_path}/{name}" if rel_path else name
            full_abs_path = os.path.join(abs_path, name)
            
            has_hash = "#" in name
            is_case_folder = re.search(r'^\d+_.*\d{2}\.\d{2}\.\d{4}', name)
            has_chat = bool(has_hash or is_case_folder)
            
            is_unread = False
            status_color = None
            if has_chat:
                if current_user_profile:
                    is_unread = check_unread_status(current_user_profile, full_rel_path)
                stats = get_case_folder_info(full_abs_path)
                if stats: status_color = stats['status_color']

            final_folders.append({
                "name": name,
                "path": full_rel_path,
                "type": "folder",
                "has_chat": has_chat,
                "is_unread": is_unread,
                "status_color": status_color,
                "mtime": folder_stats.st_mtime
            })

        # --- PROCESS FILES ---
        final_files = []
        for entry in files_to_process:
            stats = entry.stat()
            ext = os.path.splitext(entry.name)[1].lower()
            # If recursive, use the pre-calculated rel_path, otherwise build it
            full_file_path = entry.rel_path if hasattr(entry, 'rel_path') else (f"{rel_path}/{entry.name}" if rel_path else entry.name)
            
            final_files.append({
                "name": entry.name,
                "path": full_file_path,
                "parent_folder": rel_path,
                "type": "file",
                "extension": ext,
                "size": format_file_size(stats.st_size),
                "mtime": stats.st_mtime
            })

    except Exception as e:
        print(f"Error in get_folder_contents_api: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({
        "folders": final_folders,
        "files": final_files,
        "has_next": has_next,
        "saved_draft": saved_draft,
        "folder_info": folder_info,

        "auto_fill": auto_fill_data,
        "page": page
    })


def parse_folder_metadata(folder_name):
    """
    Unified Parsing Logic for Auto-Fill based on the NEW universal pattern:
    {ID}_#{NAME}#_{PRODUCT}_{LOC}_{DATE}_{SITE}_{OFF}_{REF}
    """
    print(f"DEBUG: Analyzing Folder -> {folder_name}")

    metadata = {
        'file_no': '',
        'applicant_name': '',
        'product': 'notselected'
    }
    
    # --- UPDATED PRODUCT CODE MAPPING ---
    # Left side: 4-letter code from folder name
    # Right side: The exact value="" from the feedback.html dropdown
    product_map = {
        'PRCS': '1st purchase',
        'CONS': 'construction',
        'TOPU': 'topup',
        'PDPD': 'pd',
        'RESL': 'resale',
        'RENO': 'renovation',
        'TAKE': 'takeover',
        'LAPL': 'lap',
        'EXTE': 'extension',
        'NPAN': 'npa',

        
        # (Optional) Keeping some old legacy codes just in case 
        # someone clicks an older folder created before the update
        'TOUP': 'topup', 'TOP': 'topup',
        'RESA': 'resale',
        'BLLP': 'lap', 'SBLM': 'lap', 'CCOL': 'lap', 'LAP': 'lap',
        'PUCH': '1st purchase', 'LAND': '1st purchase', 'PURC': '1st purchase',
        'CONST': 'construction',
        'PD': 'pd',
        'TAKO': 'takeover', 'HLBG': 'takeover', 'BT': 'takeover',
        'NPA': 'npa'
    }

    # Helper to find product in string
    def find_product(text):
        keys = "|".join(product_map.keys())
        match = re.search(f'(?:_|^| )({keys})(?:_|$| )', text, re.IGNORECASE)
        if match:
            return product_map.get(match.group(1).upper(), 'notselected')
        return 'notselected'

    # 1. Extract File No:
    file_match = re.match(r'^(\d+)', folder_name)
    if file_match: 
        metadata['file_no'] = file_match.group(1)

    # 2. Extract Applicant Name:
    name_match = re.search(r'#([^#]+)#', folder_name)
    if name_match: 
        metadata['applicant_name'] = name_match.group(1).replace('_', ' ').strip()
    
    # 3. Extract Product
    metadata['product'] = find_product(folder_name)

    # 4. Fallback for older folders
    if not metadata['file_no'] and '_' in folder_name:
        parts = folder_name.split('_')
        if parts[0].isdigit():
            metadata['file_no'] = parts[0]
        if not metadata['applicant_name']:
            possible_name = parts[-1].strip()
            metadata['applicant_name'] = re.sub(r'\.[a-zA-Z0-9]{3,}$', '', possible_name)

    print(f"DEBUG: Auto-fill extracted -> {metadata}")
    return metadata


def update_user_activity(user_id):
    """Updates last_seen for a user. Call this on key interactions."""
    try:
        UserProfile.objects.filter(id=user_id).update(last_seen=timezone.now())
    except Exception:
        pass


@csrf_protect
@require_POST
def auto_save_api(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    update_user_activity(user_id)

    try:
        user = UserProfile.objects.get(id=user_id)
        data = json.loads(request.body)
        
        current_folder_path = data.get('folder_path', "").strip("/")
        form_payload = data.get('payload', {})
        
        metrics = form_payload.get('completion_metrics', {})
        try:
            score = int(float(metrics.get('percent', 0)))
        except (ValueError, TypeError):
            score = 0
        
        # ---------------------------------------------------------
        # 🛑 LOGIC UPDATE 1: COMPLETION THRESHOLD (< 10%)
        # ---------------------------------------------------------
        # If less than 10%, we acknowledge the request but DO NOT save to DB.
        if score < 5:
            return JsonResponse({
                'success': True,
                'message': 'Skipped: Completion under 20%',
                'saved': False
            })

        checklist = form_payload.get('Valuers_Checklist', {})
        user_file_no = str(checklist.get('Office_file_no', "")).strip()
        applicant_name = str(checklist.get('applicant_name', "")).strip()

        if not user_file_no or not applicant_name:
             return JsonResponse({'success': False, 'error': 'Required identification missing'})

        # Initial assumption: we save to current folder
        final_target_path = current_folder_path
        folder_changed = False

        # =========================================================
        # 🛡️ STRICT MATCHING & REDIRECTION LOGIC
        # =========================================================
        if score >= 20 and user_file_no:
            current_folder_name = os.path.basename(current_folder_path)
            # Handle cases where folder might not have underscores
            folder_parts = current_folder_name.split('_')
            folder_file_no = folder_parts[0] if folder_parts else ""

            if user_file_no != folder_file_no:
                print(f"🔍 Mismatch! User: {user_file_no} vs Folder: {folder_file_no}. Searching index...")
                
                # Query the In-Memory Index
                index = get_index()
                all_folders = index.get("folders", [])
                
                # Find folder starting with "user_file_no_"
                match = next((f for f in all_folders if f['name'].startswith(f"{user_file_no}_")), None)
                
                if match:
                    final_target_path = match['path']
                    folder_changed = True
                    print(f"Found correct folder: {final_target_path}")
                else:
                    # Critical mismatch and no valid folder found
                    print(f"❌ No matching folder found for {user_file_no}. Save rejected.")
                    return JsonResponse({
                        'success': False,
                        'error': 'File Number mismatch. No valid folder found in index.',
                        'mismatch': True
                    })

        # =========================================================
        # 💾 LOGIC UPDATE 2: SAVE OPERATION (Unique Office File No)
        # =========================================================
        
        # We use update_or_create using 'office_file_no' as the ONLY lookup field.
        # This prevents duplicate records for the same file number.
        report, created = SiteVisitReport.objects.update_or_create(
            office_file_no=user_file_no,  # <--- Lookup Key
            defaults={
                'user': user,  # Update the user ownership to the last person who saved
                'target_folder': final_target_path,
                'form_data': form_payload,
                'applicant_name': applicant_name,
                'completion_score': score
            }
        )

        # Cleanup: If we moved folders, ensure no lingering drafts exist under the old folder name
        # (Optional, but good for hygiene)
        if folder_changed:
            SiteVisitReport.objects.filter(
                user=user,
                target_folder=current_folder_path
            ).exclude(id=report.id).delete()

        # 🧹 DELETE CLEARED SKETCHES: If a user cleared the canvas, delete backing records so they don't resurrect on reload
        images_data = form_payload.get('images', {})
        vectors_data = form_payload.get('vectors', {})
        for sketch_key in set(images_data.keys()) | set(vectors_data.keys()):
            if not images_data.get(sketch_key) and not vectors_data.get(sketch_key):
                ReportSketch.objects.filter(report=report, source_key=sketch_key).delete()

        return JsonResponse({
            'success': True,
            'last_saved': datetime.now().strftime("%H:%M:%S"),
            'report_id': report.id,
            'new_folder_path': final_target_path if folder_changed else None,
            'saved': True
        })

    except Exception as e:
        print(f"Auto-save error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
            
# chat folders////////////////


def check_unread_status(user_profile, folder_path):
    # Ensure we actually have a user profile before querying
    if not user_profile:
        return False

    # 1. Check if there are ANY messages in this folder
    last_msg = FolderChatMessage.objects.filter(folder_path=folder_path).last()
    if not last_msg:
        return False  # No messages, so it can't be unread
    
    # 2. Check when the user last visited
    visit = FolderChatVisit.objects.filter(user=user_profile, folder_path=folder_path).first()
    
    # 3. Logic: If never visited OR last message is newer than visit -> Unread
    if not visit:
        return True 
    
    return last_msg.timestamp > visit.last_visit


@require_http_methods(["GET"])
def serve_file(request):
    # 1. Get the relative path from the URL
    raw_path = request.GET.get('path')
    if not raw_path:
        return HttpResponseBadRequest("Missing 'path' parameter")
    
    # 2. Decode URL (converts '%20' to space)
    rel_path = unquote(raw_path)
    
    # 3. Construct the Full Path using DOCUMENTS_ROOT (G:\My Drive...)
    # We use os.path.normpath to fix slashes (forward vs backward)
    full_path = os.path.normpath(os.path.join(settings.DOCUMENTS_ROOT, rel_path))

    # 4. Debugging: Print exactly where we are looking (Check your terminal!)
    print(f"--- SERVE FILE DEBUG ---")
    print(f"Looking for: {rel_path}")
    print(f"Full Path:   {full_path}")

    # 5. Check if file exists
    if not os.path.exists(full_path):
        print("ERROR: File not found on disk.")
        return HttpResponseNotFound(f"File not found at: {full_path}")

    # 6. Security Check (Prevent accessing files outside G:\My Drive)
    # This ensures someone can't ask for "..\..\Windows\System32"
    if not full_path.startswith(os.path.normpath(settings.DOCUMENTS_ROOT)):
        return HttpResponseNotFound("Access Denied: Invalid file path.")

    # 7. Serve the file
    content_type, _ = mimetypes.guess_type(full_path)
    if not content_type:
        content_type = 'application/octet-stream'

    response = FileResponse(open(full_path, 'rb'), content_type=content_type)
    
    is_download = request.GET.get('download') == 'true'
    
    if is_download:
        # Force browser to save file
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(full_path)}"'
    elif 'pdf' in content_type:
        # Otherwise, preview it
        response['Content-Disposition'] = 'inline'  
    return response


def get_thumbnail(request):
    # 1. Get Path
    rel_path = request.GET.get('path')
    if not rel_path:
        return HttpResponseNotFound()
    
    # 2. Clean Path (Handle spaces/encoding)
    rel_path = unquote(rel_path)
    full_path = os.path.normpath(os.path.join(settings.DOCUMENTS_ROOT, rel_path))

    # 3. Security & Existence Check
    if not full_path.startswith(os.path.normpath(settings.DOCUMENTS_ROOT)):
        return HttpResponseNotFound()
    if not os.path.exists(full_path):
        return HttpResponseNotFound()

    # 4. Check Cache (Optional: To make it even faster on reload)
    # cache_key = f"thumb_{rel_path}"
    # cached_img = cache.get(cache_key)
    # if cached_img:
    #     return HttpResponse(cached_img, content_type="image/jpeg")

    try:
        # 5. Generate Image using PyMuPDF (Fast!)
        doc = fitz.open(full_path)
        page = doc.load_page(0)  # Load first page
        pix = page.get_pixmap(matrix=fitz.Matrix(0.5, 0.5))  # 50% scale (thumbnail size)
        img_data = pix.tobytes("jpg")  # Convert to JPG bytes
        
        # cache.set(cache_key, img_data, 60*60*24) # Cache for 24 hours
        return HttpResponse(img_data, content_type="image/jpeg")
        
    except Exception as e:
        print(f"Thumbnail Error: {e}")
        # Return a 1x1 pixel empty image or 404 so browser shows default icon
        return HttpResponseNotFound()

# Add this function to views.py
# It uses 'fitz' which you already imported


@require_http_methods(["GET"])
def render_pdf_page(request):
    """
    Renders a specific page of a PDF as an image (PNG).
    Usage: /coreapi/render-page/?path=Folder/file.pdf&page=5&zoom=1.5
    """
    # 1. Get Parameters
    rel_path = request.GET.get('path')
    page_num = int(request.GET.get('page', 0))  # Default to Page 0 (First page)
    zoom = float(request.GET.get('zoom', 1.0))  # Zoom level (1.0 = 100%, 2.0 = 200%)

    if not rel_path:
        return HttpResponseBadRequest("Missing path")

    # 2. Construct Path
    rel_path = unquote(rel_path)
    full_path = os.path.normpath(os.path.join(settings.DOCUMENTS_ROOT, rel_path))

    # 3. Security Check
    if not full_path.startswith(os.path.normpath(settings.DOCUMENTS_ROOT)):
        return HttpResponseNotFound("Access Denied")
    
    if not os.path.exists(full_path):
        return HttpResponseNotFound("File not found")

    try:
        # 4. Open PDF with PyMuPDF
        doc = fitz.open(full_path)
        
        # Validation
        if page_num < 0 or page_num >= len(doc):
             doc.close()
             return HttpResponseNotFound("Page number out of range")

        # 5. Render the Page
        page = doc.load_page(page_num)
        
        # Matrix handles the Zoom/Quality
        # 1.5 zoom is good balance of quality vs speed
        mat = fitz.Matrix(zoom, zoom) 
        pix = page.get_pixmap(matrix=mat, alpha=False)
        
        # 6. Convert to PNG bytes
        img_data = pix.tobytes("png")
        
        doc.close()

        # 7. Return Image
        return HttpResponse(img_data, content_type="image/png")

    except Exception as e:
        print(f"PDF Render Error: {e}")
        return HttpResponseBadRequest("Error rendering PDF")


@require_http_methods(["GET"])
def get_file_info(request):
    """Get detailed file information"""
    file_path = request.GET.get('path', '')
    
    if not file_path:
        return JsonResponse({'error': 'File path not provided'}, status=400)
    
    full_path = os.path.join(DOCUMENTS_FOLDER, file_path)
    full_path = os.path.normpath(full_path)
    
    if not full_path.startswith(DOCUMENTS_FOLDER) or not os.path.exists(full_path):
        return JsonResponse({'error': 'File not found'}, status=404)
    
    file_stat = os.stat(full_path)
    
    info = {
        'name': os.path.basename(full_path),
        'size': format_file_size(file_stat.st_size),
        'modified': file_stat.st_mtime,
        'created': file_stat.st_ctime,
        'extension': os.path.splitext(full_path)[1].lower(),
    }
    
    # If PDF, add page count
    if info['extension'] == '.pdf':
        try:
            doc = fitz.open(full_path)
            info['page_count'] = len(doc)
            doc.close()
        except Exception as e:
            print(f"Error getting PDF page count for {full_path}: {e}")
            info['page_count'] = 0

    return JsonResponse(info)


@require_http_methods(["GET"])
def list_all_folders_api(request):
    folders = []

    for root, dirs, _ in os.walk(DOCUMENTS_FOLDER):
        rel = os.path.relpath(root, DOCUMENTS_FOLDER)
        if rel != ".":
            folders.append(rel.replace("\\", "/"))

    return JsonResponse({"folders": folders})

# ------------------------------------------------


@csrf_protect
@require_http_methods(["POST"])
def analyze_file(request):
    try:
        data = json.loads(request.body)

        files = data.get("files", [])
        selected_folder = data.get("folder")

        if not selected_folder:
            return JsonResponse({"success": False, "error": "Folder not provided"})

        user_folder = os.path.join(settings.DOCUMENTS_ROOT, selected_folder)

        if not os.path.isdir(user_folder):
            return JsonResponse({"success": False, "error": "Selected folder does not exist"})

        if not files:
            return JsonResponse({"success": False, "error": "No files selected"})

        ocr_inputs = []
        direct_texts = []

        # ---------------------------
        # STEP 1: CLASSIFY FILES
        # ---------------------------
        for f in files:
            rel_path = f.get("file_path")
            if not rel_path:
                continue

            abs_path = os.path.normpath(
                os.path.join(settings.DOCUMENTS_ROOT, rel_path)
            )

            if not abs_path.startswith(settings.DOCUMENTS_ROOT):
                return JsonResponse({"success": False, "error": "Invalid file path"})

            if not os.path.exists(abs_path):
                continue

            ext = abs_path.lower()

            if ext.endswith(".pdf"):
                ocr_inputs.append(abs_path)

            elif ext.endswith((".jpg", ".jpeg", ".png", ".bmp", ".tiff")):
                img = Image.open(abs_path).convert("RGB")
                buf = io.BytesIO()
                img.save(buf, format="PDF")
                buf.seek(0)
                ocr_inputs.append(buf)

            elif ext.endswith(".docx"):
                direct_texts.append(extract_text_from_docx(abs_path))

            elif ext.endswith((".xls", ".xlsx")):
                direct_texts.append(extract_text_from_excel(abs_path))

        extracted_text = []

        # ---------------------------
        # STEP 2: OCR (ONLY IF NEEDED)
        # ---------------------------
        if ocr_inputs:
            merger = PdfMerger()
            for item in ocr_inputs:
                merger.append(item)

            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            temp_pdf_path = os.path.join(
                settings.GENERATED_PDFS_ROOT,
                f"temp_{timestamp}.pdf"
            )

            merger.write(temp_pdf_path)
            merger.close()

            from google.cloud import vision
            client = vision.ImageAnnotatorClient()

            with open(temp_pdf_path, "rb") as f:
                pdf_content = f.read()

            request_vision = vision.AnnotateFileRequest(
                input_config=vision.InputConfig(
                    content=pdf_content,
                    mime_type="application/pdf"
                ),
                features=[
                    vision.Feature(
                        type_=vision.Feature.Type.DOCUMENT_TEXT_DETECTION
                    )
                ]
            )

            response = client.batch_annotate_files(
                requests=[request_vision]
            )

            for file_response in response.responses:
                for page_response in file_response.responses:
                    if page_response.full_text_annotation:
                        extracted_text.append(
                            page_response.full_text_annotation.text
                        )

            os.remove(temp_pdf_path)

        # ---------------------------
        # STEP 3: MERGE ALL TEXT
        # ---------------------------
        full_text = "\n".join(direct_texts + extracted_text).strip()

        if not full_text:
            return JsonResponse({
                "success": False,
                "error": "No text detected"
            })

        # ---------------------------
        # STEP 4: SAVE RESULT
        # ---------------------------
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        txt_filename = f"extracted_{timestamp}.txt"
        txt_path = os.path.join(user_folder, txt_filename)

        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(full_text)

        return JsonResponse({
            "success": True,
            "message": "Text extracted successfully",
            "output_file": txt_filename
        })

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def extract_text_from_docx(path):
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_text_from_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    texts = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    texts.append(str(cell.value))

    return "\n".join(texts)


def feedback(request):
    """Render the feedback form page"""
    if not request.session.get("user_id"):
        return redirect("coreapi:login_page")
    context = {
        'google_maps_api_key': settings.GOOGLE_MAPS_API_KEY,
    }
    return render(request, "feedback.html", context)


# --- HELPER: DRAW VECTORS TO IMAGE (Server Side) ---
def generate_image_from_vectors(vector_list, width=1000, height=1000):
    """
    Takes a list of stroke objects (from frontend JSON) and draws them 
    onto a white canvas using Python's Pillow library.
    """
    if not vector_list:
        return None

    try:
        # 1. Create a white canvas
        image = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(image)

        # 2. Loop through every stroke in the vector list
        for stroke in vector_list:
            points = stroke.get('points', [])
            color = stroke.get('color', '#000000')
            width = int(stroke.get('size', 3))
            
            # Convert JSON points [{'x':1, 'y':1}, ...] to Python tuples [(1,1), ...]
            # We add an offset (e.g. +500) if your coordinates are negative/centered
            # Assuming standard top-left coordinates here, but you might need to adjust
            xy_points = [(p['x'], p['y']) for p in points]

            if len(xy_points) > 1:
                # Draw the line connecting points
                draw.line(xy_points, fill=color, width=width, joint='curve')
            elif len(xy_points) == 1:
                # Draw a dot if it's just one point
                x, y = xy_points[0]
                draw.ellipse([x - width, y - width, x + width, y + width], fill=color)

        # 3. Save to memory buffer
        output = io.BytesIO()
        image.save(output, format="PNG")
        return ContentFile(output.getvalue(), name="generated_sketch.png")

    except Exception as e:
        print(f"Vector generation error: {e}")
        return None


# --- MAIN VIEW (Refactored for Strict Folder Matching) ---
@csrf_protect
@require_POST
def save_feedback(request):
    user_id = request.session.get("user_id")
    if not user_id: 
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        user = UserProfile.objects.get(id=user_id)
        request_data = json.loads(request.body)
        
        # 1. Get Payload
        payload = request_data.get('payload', {})
        
        # 2. Extract Data
        checklist_data = payload.get('Valuers_Checklist', {})
        office_file_no_val = str(checklist_data.get('Office_file_no', "")).strip()
        applicant_name_val = str(checklist_data.get('applicant_name', "")).strip()
        current_folder_path = payload.get('target_folder', "").strip("/")
        
        # 3. Completion Threshold Check
        metrics = payload.get('completion_metrics', {})
        try:
            score = int(float(metrics.get('percent', 0)))
        except (ValueError, TypeError):
            score = 0

        final_target_path = current_folder_path
        folder_changed = False

        # =========================================================
        # TRAFFIC CONTROLLER (Strict Folder Matching)
        # =========================================================
        # Logic triggers only if form is >= 20% filled
        if score >= 20 and office_file_no_val:
            # Extract ID from current folder name (e.g., "2428_Mahesh" -> "2428")
            current_folder_name = os.path.basename(current_folder_path)
            folder_file_no = current_folder_name.split('_')[0]

            if office_file_no_val != folder_file_no:
                print(f"Mismatch! User: {office_file_no_val} vs Folder: {folder_file_no}. Searching index...")
                
                # Query the In-Memory Index (Previously built at server start)
                index = get_index()
                all_folders = index.get("folders", [])
                
                # Search for folder starting with "office_file_no_val_"
                match = next((f for f in all_folders if f['name'].startswith(f"{office_file_no_val}_")), None)
                
                if match:
                    final_target_path = match['path']
                    folder_changed = True
                    print(f"Found correct folder: {final_target_path}")
                else:
                    print(f"No matching folder found for {office_file_no_val}. Save rejected.")
                    return JsonResponse({
                        'success': False,
                        'error': f'Office File No {office_file_no_val} does not match current folder and was not found in system.',
                        'mismatch': True
                    })

        # 4. IDENTIFY REPORT
        report_id = payload.get('report_id')
        report = None

        if report_id:
            report = SiteVisitReport.objects.filter(id=report_id).first()

        # FIX: Prevent UNIQUE constraint failed by checking Office File No
        if not report and office_file_no_val:
            report = SiteVisitReport.objects.filter(office_file_no=office_file_no_val).first()

        if not report:
            # Fallback: Find by user and target folder path
            report = SiteVisitReport.objects.filter(
                user=user,
                target_folder=final_target_path
            ).order_by('-updated_at').first()

        # 5. EXTRACT IMAGES AND VECTORS
        images_data = payload.pop('images', {}) 
        vectors_data = payload.get('vectors', {})

        # 6. CREATE OR UPDATE REPORT
        if report:
            print(f"Updating report: {report.id}")
            report.form_data = payload
            report.office_file_no = office_file_no_val
            report.applicant_name = applicant_name_val
            report.target_folder = final_target_path  # Update to final path
            report.completion_score = score
            report.save()
        else:
            print("Creating NEW report")
            report = SiteVisitReport.objects.create(
                user=user,
                form_data=payload,
                office_file_no=office_file_no_val,
                applicant_name=applicant_name_val,
                target_folder=final_target_path,
                completion_score=score
            )

        # 7. PROCESS SKETCHES (With Explicit Deletion Logic)
        all_sketch_keys = set(images_data.keys()) | set(vectors_data.keys())

        for source_key in all_sketch_keys:
            image_file = None 
            is_base64 = False
            
            base64_val = images_data.get(source_key)
            vector_val = vectors_data.get(source_key)
            
            # If both values are empty, the user cleared the sketch.
            if not base64_val and not vector_val:
                print(f"Deleting sketch record for: {source_key}")
                ReportSketch.objects.filter(report=report, source_key=source_key).delete()
                continue 

            # STRATEGY A: Try Base64 Image
            if base64_val and isinstance(base64_val, str) and base64_val.startswith('data:image'):
                try:
                    format_header, imgstr = base64_val.split(';base64,') 
                    ext = format_header.split('/')[-1]
                    file_name = f"{source_key}_{report.id}.{ext}"
                    image_file = ContentFile(base64.b64decode(imgstr), name=file_name)
                    is_base64 = True
                except Exception as e:
                    print(f"Base64 error for {source_key}: {e}")

            # STRATEGY B: Fallback to Vectors
            if not image_file and vector_val: 
                print(f"Generating image from vectors for: {source_key}")
                generated_file = generate_image_from_vectors(vector_val)
                if generated_file:
                    generated_file.name = f"{source_key}_{report.id}_generated.png"
                    image_file = generated_file

            # SAVE/UPDATE only if we have a valid file
            if image_file:
                ReportSketch.objects.update_or_create(
                    report=report,
                    source_key=source_key,
                    defaults={'image': image_file}
                )

        return JsonResponse({
            'success': True,
            'report_id': report.id,
            'new_folder_path': final_target_path if folder_changed else None,
            'redirect_url': f"/coreapi/pdf-editor/{report.id}/"
        })

    except Exception as e:
        print(f"Save Feedback Error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
             
# --- STEP 2: RENDER EDITOR PAGE ---


def get_report_data_with_sketches(report):
    """
    Helper function: Merges the saved form text with the images 
    stored in the ReportSketch table.
    """
    # 1. Get the base text data
    context_data = report.form_data
    if isinstance(context_data, str):
        try:
            context_data = json.loads(context_data)
        except json.JSONDecodeError:
            context_data = {}
            
    if not context_data:
        context_data = {}

    # 2. Ensure 'images' key exists
    if 'images' not in context_data:
        context_data['images'] = {}

    # 3. Fetch images from ReportSketch table and re-inject them
    # This puts the URLs back into the JSON so the frontend sees them
    sketches = ReportSketch.objects.filter(report=report)
    for sketch in sketches:
        if sketch.image:
            # We use the file URL so the browser can load it
            context_data['images'][sketch.source_key] = sketch.image.url

    return context_data


def pdf_editor_page(request, report_id):
    """
    Renders the PDF Editor. Now includes the re-injected sketch images.
    """
    report = get_object_or_404(SiteVisitReport, id=report_id)
    
    # Use the helper to get data + images
    full_data = get_report_data_with_sketches(report)

    context = {
        'report_id': report_id,
        'data': full_data,
        'target_folder': report.target_folder,
        'google_maps_api_key': settings.GOOGLE_MAPS_API_KEY
    }
    
    return render(request, "pdf_editor.html", context)


@require_GET
def get_report_data(request, report_id):
    """
    API called by the PDF Editor JS (if it uses fetch).
    Now returns data + images.
    """
    report = get_object_or_404(SiteVisitReport, id=report_id)
    full_data = get_report_data_with_sketches(report)
    return JsonResponse(full_data)


@csrf_protect
@require_POST
def finalize_pdf(request):
    try:
        # --- CORRECT FIX: Force ProactorEventLoop for Windows ---
        # Playwright requires subprocesses, which only Proactor supports on Windows.
        if sys.platform == 'win32':
            asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
        # --------------------------------------------------------

        # 1. Parse Data
        payload = json.loads(request.body)
        report_id = payload.get('report_id')
        target_path = payload.get('target_folder', '')
        html_content = payload.get('html_content', '')

        if not html_content:
            return JsonResponse({'success': False, 'error': "No HTML content received"}, status=400)

        # 2. Determine Save Directory
        if not target_path or target_path == "/":
            save_dir = settings.DOCUMENTS_ROOT
        else:
            clean_target = target_path.lstrip('/')
            save_dir = os.path.join(settings.DOCUMENTS_ROOT, clean_target)
        
        if not os.path.exists(save_dir):
            os.makedirs(save_dir, exist_ok=True)

        # 3. Filename Logic
        checklist = payload.get('Valuers_Checklist', {})
        raw_file_no = checklist.get('Office_file_no') or 'Draft'
        raw_name = checklist.get('applicant_name') or 'Report'
        safe_file_no = str(raw_file_no).strip().replace(' ', '_').replace('/', '-')
        safe_name = str(raw_name).strip().replace(' ', '_').replace('/', '-')
        base_filename = f"{safe_file_no}_{safe_name}"
        
        counter = 1
        while True:
            pdf_filename = f"{base_filename}_site_report_{counter}.pdf"
            full_save_path = os.path.join(save_dir, pdf_filename)
            if not os.path.exists(full_save_path):
                break
            counter += 1

        # 4. GENERATE PDF
        print(f"Starting PDF generation for {pdf_filename} (HTML size: {len(html_content)} bytes)")
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(args=["--no-sandbox", "--disable-setuid-sandbox"])
                # Set a reasonable viewport to avoid layout issues
                context = browser.new_context(viewport={'width': 800, 'height': 1200})
                page = context.new_page()
                
                # Set content & wait for images (already base64 mostly)
                # Using 'load' instead of 'networkidle' for speed since images are local
                page.set_content(html_content, wait_until="load", timeout=60000) 
                
                # Small buffer for any late-loading resources or JS execution
                time.sleep(1)
                
                page.pdf(
                    path=full_save_path,
                    format="A4",
                    margin={ "top": "0", "bottom": "0", "left": "0", "right": "0" },
                    print_background=True,
                    scale=1.0,
                    prefer_css_page_size=False
                )
                browser.close()
            print("PDF successfully generated.")
        except Exception as pw_err:
            print(f"Playwright Error: {pw_err}")
            raise pw_err

        # 5. Backup Logic
        try:
            backup_dir = os.path.join(settings.BASE_DIR, 'generated_pdfs')
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir, exist_ok=True)
            backup_save_path = os.path.join(backup_dir, pdf_filename)
            shutil.copy2(full_save_path, backup_save_path)
            
            if report_id:
                try:
                    rpt = SiteVisitReport.objects.get(id=report_id)
                    rpt.generated_pdf_name = pdf_filename 
                    rpt.save()
                except:
                    pass

            return JsonResponse({'success': True, 'file_path': full_save_path})
        except Exception as copy_error:
            print(f"Warning: Could not save backup copy: {copy_error}")
        
        return JsonResponse({'success': True, 'file_path': full_save_path})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ==========================================
#  PDF GENERATION UTILS
# ==========================================


def fill_site_report_pdf(data, images_dict, target_folder_path, filename):
    """
    Fills 'Sitefeedbackform.pdf' template using PyMuPDF (fitz).
    """
    # 1. Setup Paths
    template_path = os.path.join(settings.BASE_DIR, 'static', 'pdf_templates', 'Sitefeedbackform.pdf')
    
    # Handle User Selected Folder
    if not target_folder_path or target_folder_path == "/":
        save_dir = settings.DOCUMENTS_ROOT
    else:
        # Remove leading slash if present to join correctly
        clean_target = target_folder_path.lstrip('/')
        save_dir = os.path.join(settings.DOCUMENTS_ROOT, clean_target)
    
    if not os.path.exists(save_dir):
        os.makedirs(save_dir, exist_ok=True)
        
    output_path = os.path.join(save_dir, filename)

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"PDF Template not found at: {template_path}")

    # 2. Flatten JSON
    flat_data = {}
    
    def flatten(y, prefix=""):
        for k, v in y.items():
            full_key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                flatten(v, full_key)
            elif isinstance(v, list):
                flat_data[full_key] = v 
                # Check items for checkboxes
                for item in v:
                    flat_data[f"{full_key}.{item}"] = True
            else:
                flat_data[full_key] = v
                flat_data[k] = v 

    flatten(data)

    # 3. Open PDF & Fill
    doc = fitz.open(template_path)

    for page in doc:
        widgets = list(page.widgets())
        
        for widget in widgets:
            name = widget.field_name
            
            # --- A. IMAGE INSERTION ---
            if name in images_dict:
                b64_img = images_dict[name]
                if b64_img and 'base64,' in b64_img:
                    try:
                        img_data = base64.b64decode(b64_img.split('base64,')[1])
                        rect = widget.rect
                        page.insert_image(rect, stream=img_data, keep_proportion=True)
                        page.delete_widget(widget)
                        continue 
                    except Exception as e:
                        print(f"Image error for {name}: {e}")

            # --- B. DATA FILLING ---
            
            # Checkbox Logic
            if widget.field_type == fitz.PDF_WIDGET_TYPE_CHECKBOX:
                is_checked = False
                
                # Case 1: Boolean match (e.g. data is True)
                if name in flat_data and flat_data[name] is True:
                    is_checked = True
                # Case 2: String match (e.g. data is "true")
                elif name in flat_data and str(flat_data[name]).lower() == "true":
                    is_checked = True
                # Case 3: List membership (Value inside a list)
                else:
                    for key, val in flat_data.items():
                        if isinstance(val, list) and name in val:
                            is_checked = True
                            break
                            
                if is_checked:
                    widget.field_value = True
                    widget.update()

            # Text Field Logic
            elif name in flat_data:
                val = flat_data[name]
                if val is not None and not isinstance(val, list) and not isinstance(val, dict):
                    widget.field_value = str(val)
                    widget.update()

    # 4. Save
    doc.flatten_form_fields()
    doc.save(output_path)
    doc.close()
    
    return output_path


def assetlinks(request):
    data = [{
      "relation": ["delegate_permission/common.handle_all_urls"],
      "target": {
        "namespace": "android_app",
        "package_name": "com.vadrida.app",
        "sha256_cert_fingerprints": [
            "8B:52:BA:5E:C0:CF:17:21:D7:1D:E6:60:41:E7:4F:F1:9D:4A:84:CA:4D:54:8E:77:58:3B:4D:AB:74:7E:2F:38"
        ]
      }
    }]
    return JsonResponse(data, safe=False)


@require_GET
def db_case_search_api(request):
    raw_bank = request.GET.get('bank', '')  # e.g. "1000"
    year_id = request.GET.get('year', '')  # e.g. "26"
    dist_id = request.GET.get('dist', '')  # e.g. "02"
    user_input = request.GET.get('q', '').strip()  # e.g. "1"

    # 1. Convert Bank ID to 2-digit Short Code (1000 -> 01 or 06 -> 06)
    bank_short_code = ""
    if raw_bank.isdigit():
        if len(raw_bank) <= 2:
            bank_short_code = raw_bank.zfill(2)
        else:
            bank_short_code = str(int(raw_bank) // 1000).zfill(2)

    filters = Q()

    if user_input:
        if user_input.isdigit():
            # --- NUMBER SEARCH (Exact ID Match) ---
            padded_seq = user_input.zfill(4)  # "1" -> "0001"
            full_target_id = f"{bank_short_code}{year_id}{dist_id}{padded_seq}"
            
            # Since the ID contains all the info, we ONLY search this column.
            # This prevents any DB column mismatches.
            filters = Q(unique_file_no=full_target_id)
            
            # Fallback: Just in case they type the full "0126020001" manually
            if len(user_input) > 4:
                filters = Q(unique_file_no=user_input)

        else:
            # --- NAME SEARCH (Filtered by Dropdowns) ---
            # Search for the name, but restrict it to the chosen dropdowns
            filters = Q(applicant_name__icontains=user_input)
            if year_id:
                filters &= Q(year=year_id)
            if dist_id:
                filters &= Q(district_code=dist_id)
            
            # Check both possible ways you might have saved the bank in the DB
            if raw_bank:
                filters &= (Q(bank_code=raw_bank) | Q(bank_code=bank_short_code))

    # Fetch from Database
    db_results = ClientFolder.objects.filter(filters).order_by('-created_at')[:20]

    folders = []
    for case in db_results:
        # Get relative path for the file explorer UI
        rel_path = os.path.relpath(case.full_folder_path, settings.DOCUMENTS_ROOT).replace('\\', '/')
        
        # Fetch the status dot color
        stats = get_case_folder_info(case.full_folder_path)
        
        folders.append({
            "name": os.path.basename(case.full_folder_path),
            "path": rel_path,
            "type": "folder",
            "has_chat": True,
            "status_color": stats['status_color'] if stats else 'gray'
        })

    return JsonResponse({"folders": folders})


from .models import VerificationReport


@csrf_exempt
def save_verification_data(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            file_no = data.get('file_no')
            
            if not file_no:
                return JsonResponse({'success': False, 'error': 'Missing File Number'})

            # Get the logged-in user (assuming you store username in session)
            user_name = request.session.get('user_name')
            user_profile = None
            if user_name:
                user_profile = UserProfile.objects.filter(user_name=user_name).first()

            # Create or Update the completely isolated Verification Table
            verification_record, created = VerificationReport.objects.update_or_create(
                office_file_no=file_no,
                defaults={
                    'verified_by': user_profile,
                    'inspection_date': data.get('inspection_date', ''),
                    'documents_received': data.get('documents_received', []),
                    'verification_database': {
                        'DynamicDocuments': data.get('documents_received', []),
                        'OwnersData': data.get('owners_data', []),
                        'ScheduleData': data.get('schedule_data', {}),
                        'SurveyAnalysis': data.get('survey_land_extend', {}),
                        'SurveyNotes': data.get('survey_notes', ''),
                        'ApplicantName': data.get('applicantName', ''),
                        'Product': data.get('product', ''),
                        'PersonMet': data.get('personMetAtSite', '')
                    }
                }
            )

            return JsonResponse({'success': True, 'message': 'Verification Data Saved to dedicated table!'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# --- TAT STATUS VIEWER ---

BANKS = [
    {'code': '01', 'name': 'HDFC'},
    {'code': '02', 'name': 'Muthoot'},
    {'code': '03', 'name': 'Bajaj'},
    {'code': '04', 'name': 'DCB'},
    {'code': '05', 'name': 'PNBHFL'},
    {'code': '06', 'name': 'SBI'},
    {'code': '07', 'name': 'CSB'},
    {'code': '08', 'name': 'Chola'},
    {'code': '09', 'name': 'SIB'},
    {'code': '10', 'name': 'ICICI'},
    {'code': '11', 'name': 'RBL'},
    {'code': '12', 'name': 'Chola SME'},
    {'code': '13', 'name': 'Axis Finance'},
]

DISTRICTS = [
    {'code': '01', 'name': 'TVM'},
    {'code': '02', 'name': 'Kollam'},
    {'code': '03', 'name': 'Pathanamthitta'},
    {'code': '04', 'name': 'Alappuzha'},
    {'code': '05', 'name': 'Kottayam'},
    {'code': '06', 'name': 'Idukki'},
    {'code': '07', 'name': 'Ernakulam'},
    {'code': '08', 'name': 'Thrissur'},
    {'code': '09', 'name': 'Palakkad'},
    {'code': '10', 'name': 'Malappuram'},
    {'code': '11', 'name': 'Kozhikode'},
    {'code': '12', 'name': 'Wayanad'},
    {'code': '13', 'name': 'Kannur'},
    {'code': '14', 'name': 'Kasargod'},
]

def status_viewer(request):
    if not request.session.get("user_id"):
        return redirect("coreapi:login_page")

    # Fetch all client folders
    folders = ClientFolder.objects.all().order_by('-created_at')
    
    results = []
    for f in folders:
        # Get TAT info from disk/metadata
        tat_info = get_case_folder_info(f.full_folder_path, db_created_at=f.created_at)
        
        if tat_info:
            results.append({
                'unique_file_no': f.unique_file_no,
                'applicant_name': f.applicant_name,
                'created_at': f.created_at,
                'bank_code': f.bank_code,
                'district_code': f.district_code,
                'status_color': tat_info['status_color'],
                'status_label': tat_info['status_label'],
                'tat_data': tat_info
            })
        else:
            # Fallback if folder missing on disk or error
            results.append({
                'unique_file_no': f.unique_file_no,
                'applicant_name': f.applicant_name,
                'created_at': f.created_at,
                'bank_code': f.bank_code,
                'district_code': f.district_code,
                'status_color': 'gray',
                'status_label': 'FOLDER MISSING',
                'tat_data': {
                    'download_date': f.created_at.strftime('%d/%m/%Y'),
                    'tat_date': (f.created_at + timedelta(days=3)).strftime('%d/%m/%Y'),
                    'site_report_date': '--',
                    'final_report_date': '--',
                    'days_taken': (datetime.now().replace(tzinfo=None) - f.created_at.replace(tzinfo=None)).days,
                    'days_overdue': 0
                }
            })

    context = {
        'results': results,
        'banks': BANKS,
        'districts': DISTRICTS
    }
    return render(request, "status_viewer.html", context)


@csrf_protect
@require_POST
def export_status_excel_api(request):
    try:
        data = json.loads(request.body)
        file_nos = data.get('file_nos', [])
        
        folders = ClientFolder.objects.filter(unique_file_no__in=file_nos)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Status Report"
        
        # Headers
        headers = ["File No", "Applicant", "Created Date", "Download Date", "TAT Date", "Site Rpt", "Final Rpt", "Status", "Days Taken"]
        ws.append(headers)
        
        for f in folders:
            tat = get_case_folder_info(f.full_folder_path, db_created_at=f.created_at)
            if not tat: continue
            
            ws.append([
                f.unique_file_no,
                f.applicant_name,
                f.created_at.strftime('%d-%m-%Y'),
                tat['download_date'],
                tat['tat_date'],
                tat['site_report_date'],
                tat['final_report_date'],
                tat['status_label'],
                tat['days_taken']
            ])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="Status_Report.xlsx"'
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_protect
@require_POST
def export_master_status_excel_api(request):
    try:
        data = json.loads(request.body)
        selected_bank = data.get('selected_bank')
        selected_bank_name = data.get('selected_bank_name', 'Bank')
        
        # Fetch folders for this bank
        folders = ClientFolder.objects.filter(bank_code=selected_bank).order_by('district_code', '-created_at')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{selected_bank_name} Master Report"
        
        headers = ["District", "File No", "Applicant", "Created Date", "Status", "Days Taken", "Site Rpt", "Final Rpt"]
        ws.append(headers)
        
        for f in folders:
            tat = get_case_folder_info(f.full_folder_path, db_created_at=f.created_at)
            if not tat: continue
            
            # Map district name
            dist_name = next((d['name'] for d in DISTRICTS if d['code'] == f.district_code), f.district_code)
            
            ws.append([
                dist_name,
                f.unique_file_no,
                f.applicant_name,
                f.created_at.strftime('%d-%m-%Y'),
                tat['status_label'],
                tat['days_taken'],
                tat['site_report_date'],
                tat['final_report_date']
            ])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{selected_bank_name}_Master_Report.xlsx"'
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

